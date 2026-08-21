from __future__ import annotations

import csv
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[2]
REPORT_JSON = ROOT / "reports/personal-meta26-dry-run-v2.json"
DECISIONS_CSV = ROOT / "reports/personal-meta26-decisiones-humanas-final.csv"
CORRECTIONS_CSV = ROOT / "reports/personal-meta26-correcciones-propuestas.csv"
FOCALIZACION_XLSX = ROOT / "data/focalizacion-agosto-2026.xlsx"
OUTPUT_XLSX = ROOT / "reports/REVISION_PERSONAL_META26.xlsx"

BLUE = "1F4E78"
PERSONAL_BLUE = "D9EAF7"
FOCAL_GREEN = "E2F0D9"
DIAG_YELLOW = "FFF2CC"
USER_ORANGE = "FCE4D6"
SUMMARY_DARK = "375623"
WHITE = "FFFFFF"
LIGHT_RED = "F4CCCC"
THIN_GRAY = Side(style="thin", color="D9E1F2")


def repair_text(value):
    if value is None:
        return ""
    text = str(value)
    if "Ã" in text or "�" in text:
        try:
            candidate = text.encode("latin-1").decode("utf-8")
            if candidate.count("�") <= text.count("�"):
                return candidate
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
    return text


def normalize(value):
    import unicodedata
    text = repair_text(value).upper().strip()
    text = "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")
    return re.sub(r"[^A-Z0-9]+", " ", text).strip()


def load_csv(path):
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return [{key: repair_text(value) for key, value in row.items()} for row in csv.DictReader(handle)]


def parse_candidate(value):
    parts = [repair_text(part).strip() for part in (value or "").split("|")]
    return (parts + [""] * 4)[:4]


def style_tabular_sheet(ws, widths, groups, decision_col=None):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 34
    for cell in ws[1]:
        cell.font = Font(color=WHITE, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=WHITE))
    for start, end, color in groups:
        for col in range(start, end + 1):
            ws.cell(1, col).fill = PatternFill("solid", fgColor=color)
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN_GRAY)
    if decision_col and ws.max_row >= 2:
        dv = DataValidation(
            type="list",
            formula1='"ACEPTAR_PROPUESTA,CORREGIR_MANUALMENTE,MANTENER_DATO_PERSONAL,REVISAR_DESPUES"',
            allow_blank=True,
        )
        dv.promptTitle = "Decisión requerida"
        dv.prompt = "Seleccione una opción; la celda se entrega vacía."
        dv.errorTitle = "Valor no válido"
        dv.error = "Use una de las opciones de la lista."
        dv.showInputMessage = True
        dv.showErrorMessage = True
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(decision_col)}2:{get_column_letter(decision_col)}{max(ws.max_row, 2)}")
    ws.sheet_view.showGridLines = False


def main():
    report = json.loads(REPORT_JSON.read_text(encoding="utf-8-sig"))
    decisions = load_csv(DECISIONS_CSV)
    corrections = load_csv(CORRECTIONS_CSV)
    report_by_row = {int(row["fila_origen"]): row for row in report["report_rows"]}
    corrections_by_row = {}
    for item in corrections:
        corrections_by_row.setdefault(int(item["fila"]), {})[item["campo"]] = item

    coverage_types = {
        "SEDE_EXISTE_OTRA_MODALIDAD_ADICIONAL": "VARIAS_MODALIDADES_POSIBLES",
        "INSTITUCION_REQUIERE_DECISION_HUMANA": "INSTITUCION_NO_COINCIDE",
        "SEDE_REQUIERE_DECISION_HUMANA": "SEDE_NO_COINCIDE",
        "MUNICIPIO_SIN_RESOLUCION_DETERMINISTA": "COMBINACION_NO_EXISTE",
    }
    coverage_decisions = [row for row in decisions if row["tipo_problema"] in coverage_types]
    coverage_rows = []
    for item in coverage_decisions:
        row_number = int(item["fila"])
        source = report_by_row[row_number]
        proposal = parse_candidate(item["valor_oficial_encontrado"])
        if not any(proposal):
            proposal = [
                source.get("municipio_origen") or "",
                source.get("institucion_origen") or "",
                source.get("sede_origen") or "",
                source.get("modalidad_origen") or "",
            ]
            for field, index in (("municipio", 0), ("institucion", 1), ("sede", 2), ("modalidad", 3)):
                correction = corrections_by_row.get(row_number, {}).get(field)
                if correction:
                    proposal[index] = correction["valor_propuesto"]
        options = item["opciones"].strip()
        explanation = item["motivo"].strip()
        if options:
            explanation += f" Candidatos oficiales: {options}."
        confidence = "AMBIGUA" if " || " in options or item["tipo_problema"] == "SEDE_EXISTE_OTRA_MODALIDAD_ADICIONAL" else "BAJA"
        coverage_rows.append([
            row_number, source.get("cedula") or "", source.get("nombre") or "",
            repair_text(source.get("municipio_origen")), repair_text(source.get("institucion_origen")),
            repair_text(source.get("sede_origen")), repair_text(source.get("modalidad_origen")),
            *[repair_text(value) for value in proposal], coverage_types[item["tipo_problema"]], confidence,
            explanation, item["recomendacion"], "", "", "", "", "", "",
        ])

    noncoverage_codes = {
        "FECHA_FIN_REQUERIDA_FALTANTE", "FECHA_INICIO_FALTANTE", "VALOR_CASO_ESPECIAL_FALTANTE",
        "MISMA_CEDULA_NOMBRE_DIFERENTE", "CARGO_FALTANTE", "SIN_UBICACION", "UBICACION_NO_RECONOCIDA",
        "TIPO_DOCUMENTO_NO_RECONOCIDO", "TIPO_VINCULACION_NO_MAPEADO",
    }
    problem_info = {
        "FECHA_FIN_REQUERIDA_FALTANTE": ("FECHA_FIN_FALTANTE", "Fecha fin vacía", "Fecha de finalización contractual", "Completar la fecha fin del contrato"),
        "FECHA_INICIO_FALTANTE": ("FECHA_INICIO_FALTANTE", "Fecha inicio vacía", "Fecha de inicio contractual", "Completar la fecha de inicio"),
        "VALOR_CASO_ESPECIAL_FALTANTE": ("CASO_ESPECIAL", "Método de pago: CASO_ESPECIAL; valor vacío", "Valor, vigencia desde y motivo", "Definir el valor y su vigencia"),
        "MISMA_CEDULA_NOMBRE_DIFERENTE": ("CONFLICTO_IDENTIDAD", "Nombre no coincide con la persona existente", "Nombre legal confirmado", "Reutilizar la persona existente sin duplicarla, una vez confirmado el nombre"),
        "CARGO_FALTANTE": ("CARGO_FALTANTE", "Cargo vacío", "Cargo laboral", "Completar el cargo real"),
        "SIN_UBICACION": ("UBICACION_FALTANTE", "Asignación/ubicación vacía", "Ubicación laboral", "Completar la ubicación laboral"),
        "UBICACION_NO_RECONOCIDA": ("UBICACION_NO_RECONOCIDA", "Asignación laboral no reconocida", "Ubicación válida o alias confirmado", "Definir si corresponde a una ubicación existente o nueva"),
        "TIPO_DOCUMENTO_NO_RECONOCIDO": ("TIPO_DOCUMENTO", "Tipo documento no reconocido", "Tipo de documento admitido", "Confirmar y parametrizar el tipo documental"),
        "TIPO_VINCULACION_NO_MAPEADO": ("TIPO_VINCULACION", "Tipo vinculación no mapeado", "Tipo de vinculación y contrato", "Completar el tipo de vinculación"),
    }
    identity_proposals = {x["cedula"]: repair_text(x["nombre_bd"]) for x in report.get("unique_identity_conflicts", [])}
    error_rows = []
    for source in report["report_rows"]:
        for code in source.get("problemas_bloqueantes", []):
            if code not in noncoverage_codes:
                continue
            kind, current_label, needed, proposal = problem_info[code]
            current = current_label
            if code == "UBICACION_NO_RECONOCIDA":
                current = repair_text(source.get("asignacion_laboral_origen") or source.get("ubicacion_operativa_origen"))
            elif code == "TIPO_DOCUMENTO_NO_RECONOCIDO":
                current = repair_text(source.get("tipo_documento_origen"))
                proposal = current
            elif code == "TIPO_VINCULACION_NO_MAPEADO":
                current = repair_text(source.get("tipo_vinculacion_origen")) or "VACÍO"
            elif code == "MISMA_CEDULA_NOMBRE_DIFERENTE":
                current = repair_text(source.get("nombre"))
                proposal = identity_proposals.get(str(source.get("cedula")), proposal)
            error_rows.append([
                int(source["fila_origen"]), str(source.get("cedula") or ""), repair_text(source.get("nombre")),
                kind, current, needed, proposal, "", "", "",
            ])

    wb = Workbook()
    ws = wb.active
    ws.title = "REVISAR_COBERTURA"
    ws.append([
        "FILA_XLSX_PERSONAL", "CEDULA", "NOMBRE_COMPLETO", "MUNICIPIO_PERSONAL", "INSTITUCION_PERSONAL",
        "SEDE_PERSONAL", "MODALIDAD_PERSONAL", "MUNICIPIO_FOCALIZACION", "INSTITUCION_FOCALIZACION",
        "SEDE_FOCALIZACION", "MODALIDAD_FOCALIZACION", "TIPO_DIFERENCIA", "CONFIANZA", "EXPLICACION",
        "ACCION_RECOMENDADA", "DECISION_USUARIO", "MUNICIPIO_CORRECTO", "INSTITUCION_CORRECTA",
        "SEDE_CORRECTA", "MODALIDAD_CORRECTA", "OBSERVACION_USUARIO",
    ])
    for row in sorted(coverage_rows, key=lambda x: x[0]):
        ws.append(row)
    style_tabular_sheet(ws, [18, 16, 30, 20, 38, 34, 19, 20, 40, 36, 21, 28, 12, 65, 48, 26, 22, 40, 36, 22, 38], [(1, 3, BLUE), (4, 7, "4472C4"), (8, 11, "70AD47"), (12, 15, "BF9000"), (16, 21, "C65911")], 16)
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 2).number_format = "@"
        for col in range(4, 8): ws.cell(row, col).fill = PatternFill("solid", fgColor=PERSONAL_BLUE)
        for col in range(8, 12): ws.cell(row, col).fill = PatternFill("solid", fgColor=FOCAL_GREEN)
        for col in range(12, 16): ws.cell(row, col).fill = PatternFill("solid", fgColor=DIAG_YELLOW)
        for col in range(16, 22): ws.cell(row, col).fill = PatternFill("solid", fgColor=USER_ORANGE)
    ws.conditional_formatting.add(f"A2:U{ws.max_row}", FormulaRule(formula=["$M2=\"AMBIGUA\""], fill=PatternFill("solid", fgColor=LIGHT_RED)))

    errors = wb.create_sheet("ERRORES_DATOS_PERSONAL")
    errors.append(["FILA_XLSX", "CEDULA", "NOMBRE", "TIPO_PROBLEMA", "VALOR_ACTUAL", "DATO_NECESARIO", "PROPUESTA_EMPIRIA", "DECISION_USUARIO", "VALOR_CORRECTO", "OBSERVACION_USUARIO"])
    for row in sorted(error_rows, key=lambda x: (x[0], x[3])):
        errors.append(row)
    style_tabular_sheet(errors, [14, 16, 30, 28, 38, 34, 48, 26, 34, 42], [(1, 3, BLUE), (4, 7, "BF9000"), (8, 10, "C65911")], 8)
    for row in range(2, errors.max_row + 1):
        errors.cell(row, 2).number_format = "@"
        for col in range(8, 11): errors.cell(row, col).fill = PatternFill("solid", fgColor=USER_ORANGE)

    summary = wb.create_sheet("RESUMEN")
    summary.sheet_view.showGridLines = False
    summary.column_dimensions["A"].width = 45
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 55
    summary.merge_cells("A1:C1")
    summary["A1"] = "REVISIÓN PERSONAL META 26"
    summary["A1"].font = Font(color=WHITE, bold=True, size=16)
    summary["A1"].fill = PatternFill("solid", fgColor=SUMMARY_DARK)
    summary["A1"].alignment = Alignment(horizontal="center")
    manipuladoras = sum(1 for row in report["report_rows"] if "MANIPULADOR" in normalize(row.get("cargo_resuelto") or row.get("cargo_origen")))
    pending = len(coverage_rows)
    auto = manipuladoras - pending
    type_counts = Counter(row[11] for row in coverage_rows)
    error_counts = Counter(row[3] for row in error_rows)
    metrics = [
        ("Total filas Personal", len(report["report_rows"]), "Total de registros del archivo fuente"),
        ("Manipuladoras", manipuladoras, "Personas cuyo cargo es manipulador(a) de alimentos"),
        ("Filas aceptadas automáticamente", len(report["report_rows"]) - len(coverage_rows) - len({r[0] for r in error_rows}), "Sin revisión manual de cobertura ni error fuente bloqueante"),
        ("Filas enviadas a revisión manual", len(set([r[0] for r in coverage_rows] + [r[0] for r in error_rows])), "Personas únicas presentes en alguna hoja de revisión"),
        ("Modalidad distinta", type_counts["MODALIDAD_DISTINTA"], ""),
        ("Institución", sum(v for k, v in type_counts.items() if "INSTITUCION" in k), ""),
        ("Sede", sum(v for k, v in type_counts.items() if "SEDE" in k), ""),
        ("Combinación sede-modalidad", type_counts["MODALIDAD_DISTINTA"] + type_counts["VARIAS_MODALIDADES_POSIBLES"], "Sede encontrada, pero modalidad distinta o múltiple"),
        ("Ambigüedad", sum(1 for row in coverage_rows if row[12] == "AMBIGUA"), "Múltiples candidatos o modalidad no determinable"),
        ("Fechas", sum(v for k, v in error_counts.items() if "FECHA" in k), ""),
        ("Identidad", error_counts["CONFLICTO_IDENTIDAD"], ""),
        ("CASO_ESPECIAL", error_counts["CASO_ESPECIAL"], ""),
        ("Ubicación", sum(v for k, v in error_counts.items() if "UBICACION" in k), ""),
        ("Catálogo", error_counts["TIPO_DOCUMENTO"] + error_counts["TIPO_VINCULACION"], "Tipo documento o vinculación"),
        ("Cobertura requerida", 662, "Cupos oficiales requeridos"),
        ("Manipuladoras asignables automáticamente", auto, "Aceptables contra focalización sin decisión humana"),
        ("Manipuladoras pendientes de revisión", pending, "Incluidas en REVISAR_COBERTURA"),
        ("Déficit provisional", max(0, 662 - auto), "PROVISIONAL mientras existan personas pendientes de revisión"),
        ("Exceso provisional", max(0, auto - 662), "PROVISIONAL mientras existan personas pendientes de revisión"),
    ]
    summary.append(["INDICADOR", "CANTIDAD", "NOTA"])
    for row in metrics: summary.append(row)
    for cell in summary[2]:
        cell.fill = PatternFill("solid", fgColor=BLUE); cell.font = Font(color=WHITE, bold=True)
    for row in summary.iter_rows(min_row=3, max_row=summary.max_row):
        for cell in row: cell.border = Border(bottom=THIN_GRAY); cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in range(17, 22):
        for cell in summary[row]: cell.fill = PatternFill("solid", fgColor=FOCAL_GREEN)
    summary.freeze_panes = "A3"
    summary.auto_filter.ref = f"A2:C{summary.max_row}"

    ref = wb.create_sheet("REFERENCIA_FOCALIZACION")
    ref.append(["MUNICIPIO", "INSTITUCION", "SEDE", "MODALIDAD", "FOCALIZACION", "COBERTURA_REQUERIDA"])
    preview_map = {(normalize(x["municipio"]), normalize(x["institucion"]), normalize(x["sede"]), normalize(x["modalidad"])): x["requeridas"] for x in report["coverage_preview"]}
    source_wb = load_workbook(FOCALIZACION_XLSX, read_only=True, data_only=True)
    source_ws = source_wb[source_wb.sheetnames[0]]
    for values in source_ws.iter_rows(min_row=3, values_only=True):
        municipio, institution, sede, modality = [repair_text(values[i]) for i in (2, 3, 4, 5)]
        if not any((municipio, institution, sede, modality)): continue
        key = tuple(normalize(x) for x in (municipio, institution, sede, modality))
        ref.append([municipio, institution, sede, modality, values[11] or 0, preview_map.get(key, 0)])
    source_wb.close()
    style_tabular_sheet(ref, [22, 48, 46, 20, 18, 22], [(1, 6, "70AD47")])
    ref.freeze_panes = "A2"

    for ws_item in wb.worksheets:
        ws_item.sheet_properties.pageSetUpPr.fitToPage = True
        ws_item.page_setup.fitToWidth = 1
        ws_item.page_setup.fitToHeight = 0
        ws_item.sheet_properties.tabColor = {"REVISAR_COBERTURA": "4472C4", "ERRORES_DATOS_PERSONAL": "C65911", "RESUMEN": "70AD47", "REFERENCIA_FOCALIZACION": "A5A5A5"}[ws_item.title]

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(json.dumps({"output": str(OUTPUT_XLSX), "coverage_review": len(coverage_rows), "data_errors": len(error_rows), "automatic_manipuladoras": auto}, ensure_ascii=False))


if __name__ == "__main__":
    main()

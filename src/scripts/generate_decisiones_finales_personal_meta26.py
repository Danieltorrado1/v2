from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[2]
V4 = ROOT / "reports/personal-meta26-dry-run-v4.json"
SOURCE = ROOT / "data/Importacion_Personal_CONSORCIO_PAE_META_26.xlsx"
OUTPUT = ROOT / "reports/DECISIONES_FINALES_PERSONAL_META26.xlsx"
AUDIT_OUTPUT = ROOT / "reports/personal-meta26-cierre-pendientes.json"
HEADERS = ["FILA_XLSX", "CEDULA", "NOMBRE", "PROBLEMA", "VALOR_ACTUAL", "CONTEXTO", "PROPUESTA_EMPIRIA", "DECISION_USUARIO", "VALOR_USUARIO", "OBSERVACION_USUARIO"]
COLORS = {"FECHAS": "C65911", "IDENTIDADES": "7030A0", "CASOS_ESPECIALES": "BF9000", "UBICACIONES_CARGOS": "1F4E78", "CATALOGOS": "375623", "RESUMEN": "44546A"}
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9E1F2")


def clean(value):
    return "" if value is None else str(value)


def source_rows():
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["IMPORTACION_META"]
    headers = [cell.value for cell in ws[1]]
    result = {}
    for excel_row, cells in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        result[excel_row] = dict(zip(headers, cells))
    wb.close()
    return result


def context(row, src):
    parts = [
        f"Cargo: {clean(row.get('cargo_origen')) or 'VACÍO'}",
        f"Vinculación: {clean(row.get('tipo_vinculacion_origen')) or 'VACÍA'}",
        f"Contrato: {clean(row.get('tipo_contrato_origen')) or 'VACÍO'}",
        f"Ubicación/asignación: {clean(src.get('UBICACION_OPERATIVA') or src.get('ASIGNACION_LABORAL')) or 'VACÍA'}",
        f"Método pago: {clean(row.get('metodo_pago_origen')) or 'VACÍO'}",
        f"Perfil licitación: {clean(row.get('licitacion_perfil_resuelto')) or clean(src.get('PERFIL_LICITACION')) or 'NO'}",
    ]
    if row.get("municipio_origen"):
        parts.append(f"Cobertura: {row.get('municipio_origen')} | {row.get('institucion_origen')} | {row.get('sede_origen')} | {row.get('modalidad_origen')}")
    return " | ".join(parts)


def base_row(row, src, problem, current, proposal):
    return [row["fila_origen"], clean(row.get("cedula")), clean(row.get("nombre")), problem, current, context(row, src), proposal, "", "", ""]


def add_decision_sheet(wb, title, rows, options):
    ws = wb.create_sheet(title)
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=COLORS[title])
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [13, 17, 31, 34, 34, 70, 62, 29, 34, 46]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
        row[1].number_format = "@"
        for index in (7, 8, 9):
            row[index].fill = PatternFill("solid", fgColor="FFF2CC")
    if rows:
        validation = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
        validation.error = "Seleccione una opción de la lista."
        validation.showErrorMessage = True
        ws.add_data_validation(validation)
        validation.add(f"H2:H{ws.max_row}")
    return ws


def main():
    report = json.loads(V4.read_text(encoding="utf-8"))
    sources = source_rows()
    pending = [row for row in report["report_rows"] if row["categoria_final"] == "REVISAR"]
    by_fila = {row["fila_origen"]: row for row in pending}

    fechas = []
    retirement_rows = [row for row in pending if row.get("subtipo_retiro")]
    for row in retirement_rows:
        src = sources[row["fila_origen"]]
        current = f"Inicio: {clean(src.get('FECHA_INICIO_CONTRATO')) or 'VACÍO'} | Fin: {clean(src.get('FECHA_FIN_CONTRATO')) or 'VACÍO'} | Retiro: SIN FECHA"
        proposal = "El retiro ya está confirmado. Ingrese únicamente la fecha efectiva real de retiro/cierre; debe ser anterior a agosto de 2026 y no se inventará."
        fechas.append(base_row(row, src, "FECHA_RETIRO_REQUERIDA", current, proposal))
    for row in pending:
        if row.get("subtipo_retiro") or "FECHA_PENDIENTE" not in row.get("pendientes_finales", []):
            continue
        src = sources[row["fila_origen"]]
        if not src.get("FECHA_INICIO_CONTRATO"):
            fechas.append(base_row(row, src, "FECHA_INICIO_REQUERIDA", "FECHA_INICIO_CONTRATO vacía", "Ingrese la fecha real de inicio contractual."))
        elif clean(src.get("TIPO_CONTRATO")).upper() in {"TÉRMINO FIJO", "TERMINO FIJO"} and not src.get("FECHA_FIN_CONTRATO"):
            fechas.append(base_row(row, src, "FECHA_FIN_TERMINO_FIJO_REQUERIDA", f"Inicio: {src.get('FECHA_INICIO_CONTRATO')} | Fin vacía", "Ingrese la fecha fin real del contrato a término fijo."))

    identidades = []
    identity_details = {
        27: ("XLSX: LUZ MYRIAM SACRISTAN CIFUENTES | BD: LUZ MYRIAN SACRISTAN CINFUENTES | mismo documento 40430665", "Seleccione MISMA_PERSONA si las diferencias MYRIAM/MYRIAN y CIFUENTES/CINFUENTES son errores ortográficos; de lo contrario PERSONA_DISTINTA."),
        740: ("XLSX: JUAN PABLO BAENA TRUJILLO | BD: JUAN PABLO TRUJILLO BAENA | mismo documento 8780626 | vinculación 24 activa como REPRESENTANTE LEGAL", "Seleccione MISMA_PERSONA si BAENA/TRUJILLO está invertido en el XLSX; de lo contrario PERSONA_DISTINTA."),
    }
    for fila, (current, proposal) in identity_details.items():
        row, src = by_fila[fila], sources[fila]
        identidades.append(base_row(row, src, "CONFLICTO_IDENTIDAD", current, proposal))
    for fila in (772, 773):
        row, src = by_fila[fila], sources[fila]
        current = f"Documento: {row.get('cedula')} | nombres y apellidos vacíos | nacimiento: {clean(src.get('FECHA_NACIMIENTO'))} | correo: {clean(src.get('CORREO')) or 'VACÍO'}"
        identidades.append(base_row(row, src, "NOMBRE_LEGAL_FALTANTE", current, "Ingrese nombres y apellidos legales. No se crearán personas sin nombre."))

    casos = []
    for fila in (20, 181, 324, 518):
        row, src = by_fila[fila], sources[fila]
        current = f"Valor: VACÍO | Vigencia sugerible desde inicio contractual: {clean(src.get('FECHA_INICIO_CONTRATO'))} | Motivo: VACÍO"
        proposal = f"Complete valor y motivo; confirme vigencia_desde {clean(src.get('FECHA_INICIO_CONTRATO'))}. No existe valor económico en el XLSX ni estructura histórica reutilizable en BD."
        casos.append(base_row(row, src, "CASO_ESPECIAL_VALOR_VIGENCIA_MOTIVO", current, proposal))

    ubicaciones = []
    auto_cargo = {690: "ADMINISTRATIVO (perfil COORD_SUMINISTRO -> cargo equivalente 38)", 692: "ADMINISTRATIVO (perfil SUP_CALIDAD -> cargo equivalente 38)"}
    for fila in (690, 692, 725, 740, 741, 743, 764, 766, 772):
        row, src = by_fila[fila], sources[fila]
        raw_location = clean(src.get("UBICACION_OPERATIVA") or src.get("ASIGNACION_LABORAL")) or "VACÍA"
        classification = {
            690: "Cargo resuelto automáticamente por perfil; ubicación real ausente.",
            692: "Cargo resuelto automáticamente por perfil; ubicación real ausente.",
            725: "AUXILIAR DE SERVICIOS GENERALES es función/cargo, no ubicación del catálogo.",
            740: "REPRESENTANTE LEGAL es cargo; ya existe como cargo de su vinculación 24, no es ubicación.",
            741: "SEGURIDAD Y SALUD EN EL TRABAJO es área/función; no existe equivalencia inequívoca en ubicaciones.",
            743: "AUXILIAR ADMINISTRATIVO es función/cargo, no ubicación.",
            764: "AUXILIAR DE SERVICIOS GENERALES es función/cargo, no ubicación.",
            766: "COORDINADOR AUXILIAR OPERATIVO es función/cargo, no ubicación.",
            772: "OPERARIO DE BODEGA, AUXILIARES Y TRANSPORTADORES describe cargo; BODEGA es ambigua entre RI, RP y GRANADA.",
        }[fila]
        current = f"Valor fuente: {raw_location} | {auto_cargo.get(fila, '')}".strip(" |")
        proposal = classification + " Seleccione la ubicación laboral real del catálogo o indique que falta parametrización funcional."
        ubicaciones.append(base_row(row, src, "UBICACION_LABORAL_REQUERIDA", current, proposal))

    catalogos = []
    for fila in (13, 21):
        row, src = by_fila[fila], sources[fila]
        current = "Usuario indicó SEDE ENRIQUE DANIELS | CAJU-RI; focalizacion_final solo contiene esa sede con CAJM/JT-RI."
        proposal = "ACEPTAR_PROPUESTA_FOCALIZACION usaría SEDE ENRIQUE DANIELS | CAJM/JT-RI. Si no corresponde, indique la combinación personal correcta sin modificar Cobertura."
        catalogos.append(base_row(row, src, "COMBINACION_SEDE_MODALIDAD_INVALIDA", current, proposal))
    row, src = by_fila[715], sources[715]
    catalogos.append(base_row(row, src, "TIPO_DOCUMENTO_PPT_NO_CATALOGADO", "PPT | documento 6038597 | catálogo personal actual: únicamente CEDULA", "PPT en Colombia corresponde normalmente a Permiso por Protección Temporal, no a PASAPORTE. Seleccione CREAR_TIPO_PPT o corrija el tipo según documento soporte."))
    row, src = by_fila[773], sources[773]
    catalogos.append(base_row(row, src, "TIPO_VINCULACION_REQUERIDO", "Tipo vinculación y tipo contrato vacíos", "Cargo ADMINISTRATIVO, GESTIÓN DE ZONA y ASISTENCIA no prueban si es LABORAL u OPS. Seleccione el vínculo real; si es LABORAL, indique además OL, TF o TI."))

    wb = Workbook()
    wb.remove(wb.active)
    add_decision_sheet(wb, "FECHAS", fechas, ["PROPORCIONAR_FECHA", "REVISAR_SOPORTE"])
    add_decision_sheet(wb, "IDENTIDADES", identidades, ["MISMA_PERSONA", "PERSONA_DISTINTA", "COMPLETAR_NOMBRE"])
    add_decision_sheet(wb, "CASOS_ESPECIALES", casos, ["COMPLETAR_DATOS", "REVISAR_SOPORTE"])
    add_decision_sheet(wb, "UBICACIONES_CARGOS", ubicaciones, ["ASIGNAR_UBICACION_CATALOGO", "SOLICITAR_NUEVA_PARAMETRIZACION", "REVISAR_SOPORTE"])
    add_decision_sheet(wb, "CATALOGOS", catalogos, ["ACEPTAR_PROPUESTA", "CORREGIR_DATO", "CREAR_TIPO_PPT", "LABORAL_OL", "LABORAL_TF", "LABORAL_TI", "OPS", "REVISAR_SOPORTE"])

    summary = wb.create_sheet("RESUMEN")
    summary.append(["CIERRE DE PENDIENTES PERSONAL META-26", "RESULTADO"])
    metrics = [
        ("REVISAR V4", 50),
        ("Incidencias V4", 53),
        ("Incidencias adicionales detectadas: nombres faltantes", 2),
        ("TOTAL_INCIDENCIAS_AUDITADAS", 55),
        ("Incidencias resueltas automáticamente", 2),
        ("Filas resueltas completamente", 0),
        ("Filas que requieren usuario", 50),
        ("FECHAS", len(fechas)),
        ("IDENTIDADES", len(identidades)),
        ("CASOS_ESPECIALES", len(casos)),
        ("UBICACIONES_CARGOS", len(ubicaciones)),
        ("CATALOGOS", len(catalogos)),
        ("LISTA_IMPORTAR_ACTIVA", 656),
        ("LISTA_IMPORTAR_SIN_COBERTURA", 66),
        ("RETIRO_CONFIRMADO", 0),
        ("REVISAR", 50),
        ("TOTAL", 772),
        ("DUPLICADAS_ENTRE_CATEGORIAS", 0),
        ("Cobertura modificada", "NO"),
        ("BD modificada", "NO"),
    ]
    for item in metrics:
        summary.append(item)
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = summary.dimensions
    summary.column_dimensions["A"].width = 55
    summary.column_dimensions["B"].width = 28
    for cell in summary[1]:
        cell.fill = PatternFill("solid", fgColor=COLORS["RESUMEN"])
        cell.font = Font(color=WHITE, bold=True)
    for row in summary.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=THIN)
    summary.sheet_view.showGridLines = False

    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    audit = {
        "revisar_v4": 50,
        "total_incidencias_v4": 53,
        "incidencias_adicionales_nombres_faltantes": 2,
        "total_incidencias_auditadas": 55,
        "incidencias_resueltas_automaticamente": 2,
        "resoluciones_automaticas": [
            {"fila": 690, "campo": "CARGO", "valor": "ADMINISTRATIVO", "evidencia": "COORD_SUMINISTRO -> contrato_cargo_equivalente_id 38"},
            {"fila": 692, "campo": "CARGO", "valor": "ADMINISTRATIVO", "evidencia": "SUP_CALIDAD -> contrato_cargo_equivalente_id 38"},
        ],
        "filas_resueltas_automaticamente": 0,
        "requieren_usuario": 50,
        "fechas_retiro_requeridas": 18,
        "otras_fechas_requeridas": 14,
        "identidades": 4,
        "casos_especiales": 4,
        "ubicaciones": 9,
        "cargos_pendientes": 0,
        "catalogos": 4,
        "categorias": {"LISTA_IMPORTAR_ACTIVA": 656, "LISTA_IMPORTAR_SIN_COBERTURA": 66, "RETIRO_CONFIRMADO": 0, "REVISAR": 50},
        "duplicadas_entre_categorias": 0,
        "cobertura_modificada": False,
        "bd_modificada": False,
        "excel": str(OUTPUT.resolve()),
    }
    AUDIT_OUTPUT.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

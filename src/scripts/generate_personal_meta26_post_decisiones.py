from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "reports/personal-meta26-dry-run-v4.json"
OUTPUT = ROOT / "reports/REVISION_PERSONAL_META26_POST_DECISIONES.xlsx"
BLUE, GREEN, ORANGE, RED, GRAY, WHITE = "1F4E78", "548235", "C65911", "C00000", "666666", "FFFFFF"
THIN = Side(style="thin", color="D9E1F2")


def scalar(value):
    if value is None:
        return ""
    if isinstance(value, list):
        return " | ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return value


def add_table(wb, title, headers, rows, color=BLUE, widths=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append([scalar(item) for item in row])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
    for index in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(index)].width = (widths[index - 1] if widths else 22)
    return ws


def main():
    report = json.loads(INPUT.read_text(encoding="utf-8"))
    rows = report["report_rows"]
    summary = report["v4_summary"]
    wb = Workbook()
    wb.remove(wb.active)

    metrics = [
        ("Archivo de decisiones", summary["decisions_file"]),
        ("SHA-256 decisiones", summary["decisions_sha256"]),
        ("Decisiones humanas encontradas", summary["decisiones_humanas_encontradas"]),
        ("Decisiones cobertura consumidas", summary["decisiones_cobertura_consumidas"]),
        ("Decisiones datos consumidas", summary["decisiones_datos_consumidas"]),
        ("Total filas", summary["total_filas"]),
        *[(key, value) for key, value in summary["categorias"].items()],
        ("RETIRO_CONFIRMADO_FECHA_PENDIENTE", summary["retiro_confirmado_fecha_pendiente"]),
        ("Reducción desde 307 REVISAR", summary["reduccion_revisar"]),
        ("Manipuladoras XLSX", summary["manipuladoras_xlsx"]),
        ("Manipuladoras activas", summary["manipuladoras_activas"]),
        ("Manipuladoras retiradas", summary["manipuladoras_retiradas"]),
        ("Manipuladoras asignables", summary["manipuladoras_asignables"]),
        ("Manipuladoras pendientes activas", summary["manipuladoras_pendientes"]),
        ("Cobertura requerida", summary["coverage"]["requeridas_total"]),
        ("Cobertura asignada propuesta", summary["coverage"]["asignadas_total"]),
        ("Déficit", summary["coverage"]["deficit_total"]),
        ("Exceso", summary["coverage"]["exceso_total"]),
        ("Sede-modalidades completas", summary["coverage"]["completas"]),
        ("Deficitarias", summary["coverage"]["deficitarias"]),
        ("Con exceso", summary["coverage"]["excesos"]),
        ("Sin personal", summary["coverage"]["sin_personal"]),
        ("Duplicadas entre categorías", summary["duplicadas_entre_categorias"]),
        ("Escrituras BD", summary["escrituras_bd"]),
        ("Seguro smoke real", "SÍ" if summary["seguro_smoke_real"] else "NO"),
        ("Bloqueadores", summary["blockers"]),
    ]
    add_table(wb, "RESUMEN", ["INDICADOR", "RESULTADO"], metrics, GREEN, [44, 90])

    plan_headers = ["FILA_XLSX", "CEDULA", "NOMBRE", "CATEGORIA_FINAL", "PENDIENTES", "PERSONA_PLAN", "VINCULACION_PLAN", "MUNICIPIO_PROPUESTO", "INSTITUCION_PROPUESTA", "SEDE_PROPUESTA", "MODALIDAD_PROPUESTA", "DECISION_USUARIO", "OBSERVACION_USUARIO"]
    plan_rows = [[r.get(k) for k in ["fila_origen", "cedula", "nombre", "categoria_final", "pendientes_finales", "persona_plan", "vinculacion_plan", "municipio_propuesto", "institucion_propuesta", "sede_propuesta", "modalidad_propuesta", "decision_cobertura_usuario", "observacion_usuario"]] for r in rows]
    plan = add_table(wb, "PLAN_772", plan_headers, plan_rows, BLUE, [13, 17, 32, 30, 32, 21, 24, 20, 42, 40, 20, 26, 50])
    for row in range(2, plan.max_row + 1):
        plan.cell(row, 2).number_format = "@"
        category = plan.cell(row, 4).value
        color = "E2F0D9" if category.startswith("LISTA") else "F4CCCC" if category == "REVISAR" else "FFF2CC"
        plan.cell(row, 4).fill = PatternFill("solid", fgColor=color)

    consumed = [r for r in rows if r.get("decision_cobertura_usuario") or r.get("observacion_usuario")]
    add_table(wb, "DECISIONES_CONSUMIDAS", ["FILA", "CEDULA", "NOMBRE", "DECISION", "OBSERVACION", "VALIDADA", "FUENTE", "RESULTADO"], [[r["fila_origen"], r["cedula"], r["nombre"], r.get("decision_cobertura_usuario"), r.get("observacion_usuario"), r.get("decision_validada"), r.get("decision_fuente"), r.get("categoria_final")] for r in consumed], GREEN, [12, 17, 32, 27, 52, 14, 25, 28])

    retired = [r for r in rows if r.get("subtipo_retiro")]
    add_table(wb, "RETIROS", ["FILA", "CEDULA", "NOMBRE", "MUNICIPIO", "INSTITUCION_ANTERIOR", "SEDE_ANTERIOR", "FECHA_INICIO_XLSX", "FECHA_FIN_XLSX", "FECHA_RETIRO_DISPONIBLE", "PERSONA_EXISTENTE", "VINCULACION_EXISTENTE", "CLASIFICACION", "OBSERVACION_USUARIO"], [[r.get(k) for k in ["fila_origen", "cedula", "nombre", "municipio_origen", "institucion_origen", "sede_origen", "fecha_inicio_xlsx", "fecha_fin_xlsx", "fecha_retiro_disponible", "retiro_persona_estado", "retiro_vinculacion_estado", "subtipo_retiro", "observacion_usuario"]] for r in retired], ORANGE, [12, 17, 31, 20, 42, 40, 20, 18, 24, 20, 24, 40, 52])

    pending = [r for r in rows if r["categoria_final"] == "REVISAR"]
    add_table(wb, "PENDIENTES", ["FILA", "CEDULA", "NOMBRE", "TIPOS_PENDIENTES", "SUBTIPO_RETIRO", "PROBLEMAS_ORIGINALES", "DECISION_USUARIO", "OBSERVACION", "ACCION"] , [[r["fila_origen"], r["cedula"], r["nombre"], r["pendientes_finales"], r.get("subtipo_retiro"), r["problemas_bloqueantes"], r.get("decision_cobertura_usuario"), r.get("observacion_usuario"), "Completar dato real y volver a ejecutar dry-run; no inventar ni aplicar automáticamente"] for r in pending], RED, [12, 17, 32, 36, 42, 45, 26, 55, 65])

    coverage = report["coverage_preview"]
    add_table(wb, "COBERTURA_PROPUESTA", ["MUNICIPIO", "INSTITUCION", "SEDE", "MODALIDAD", "REQUERIDAS", "ASIGNADAS", "DIFERENCIA", "ESTADO"], [[r.get(k) for k in ["municipio", "institucion", "sede", "modalidad", "requeridas", "asignadas_propuestas", "diferencia", "estado"]] for r in coverage], GREEN, [21, 45, 43, 21, 16, 16, 16, 18])
    lic = summary["licitacion"]["perfiles"]
    add_table(wb, "LICITACION", ["PERFIL", "REQUERIDOS", "PRESENTADOS_VALIDOS", "DIFERENCIA", "ESTADO"], [[r.get(k) for k in ["perfil", "requeridos", "presentados", "diferencia", "estado"]] for r in lic], GRAY, [32, 18, 23, 18, 18])

    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
    wb.save(OUTPUT)
    print(json.dumps({"output": str(OUTPUT.resolve()), "sheets": wb.sheetnames, "rows_plan": len(plan_rows), "pending": len(pending), "retirements": len(retired)}, ensure_ascii=False))


if __name__ == "__main__":
    main()

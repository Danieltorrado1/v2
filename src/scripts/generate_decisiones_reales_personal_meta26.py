from __future__ import annotations

from pathlib import Path
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / 'reports' / 'DECISIONES_FINALES_PERSONAL_META26.xlsx'
OUTPUT = ROOT / 'reports' / 'DECISIONES_REALES_PERSONAL_META26.xlsx'
HEADERS = [
    'FILA_XLSX', 'CEDULA', 'NOMBRE', 'PROBLEMA', 'VALOR_ACTUAL',
    'CONTEXTO', 'PROPUESTA_EMPIRIA', 'DECISION_USUARIO', 'VALOR_USUARIO', 'OBSERVACION_USUARIO'
]
COLORS = {
    'FECHAS': 'C65911',
    'IDENTIDADES': '7030A0',
    'CASOS_ESPECIALES': 'BF9000',
    'UBICACIONES': '1F4E78',
    'CATALOGOS': '375623',
    'RESUMEN': '44546A',
}
WHITE = 'FFFFFF'
INPUT_FILL = 'FFF2CC'
THIN = Side(style='thin', color='D9E1F2')


def clean(value):
    return '' if value is None else str(value)


def normalize_text(value: str) -> str:
    return ' '.join(clean(value).strip().upper().split())


def to_iso_date(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, (int, float)):
        return from_excel(value).strftime('%Y-%m-%d')
    text = clean(value).strip()
    if not text:
        return None
    if text.count('-') == 2 and len(text) == 10:
        return text
    if text.count('/') == 2:
        day, month, year = text.split('/')
        if len(year) == 4:
            return f'{year}-{month.zfill(2)}-{day.zfill(2)}'
    return None


def read_sheet_rows(workbook, sheet_name):
    ws = workbook[sheet_name]
    headers = [clean(cell.value) for cell in ws[1]]
    rows = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        data = dict(zip(headers, values))
        if not clean(data.get('FILA_XLSX')).strip():
            continue
        rows.append({'_source_row': excel_row, **data})
    return rows


def build_fecha_rows(rows):
    filtered = []
    for row in rows:
        fila = int(row['FILA_XLSX'])
        problema = clean(row.get('PROBLEMA'))
        iso = to_iso_date(row.get('VALOR_USUARIO'))
        obs = normalize_text(clean(row.get('OBSERVACION_USUARIO')))
        if problema == 'FECHA_RETIRO_REQUERIDA' and iso == '2026-08-02':
            continue
        if fila == 346 and 'NO ESTA RETIRADA' in obs and 'CONTINUA LABORANDO' in obs:
            continue
        if fila == 74 and iso == '2026-08-13':
            continue
        filtered.append({
            **row,
            'DECISION_USUARIO': '',
            'VALOR_USUARIO': '',
            'OBSERVACION_USUARIO': '',
            'PROPUESTA_EMPIRIA': (
                'Diligencie la fecha exacta requerida. '
                + ('No diligencie retiro: los 17 retiros del 2026-08-02 ya quedaron aceptados.' if problema == 'FECHA_RETIRO_REQUERIDA' else '')
            ).strip(),
        })
    return filtered


def build_identity_rows(rows):
    result = []
    for row in rows:
        fila = int(row['FILA_XLSX'])
        proposal = clean(row.get('PROPUESTA_EMPIRIA'))
        if fila in (772, 773):
            proposal = 'Diligencie NOMBRE_LEGAL_COMPLETO exacto en VALOR_USUARIO.'
        result.append({
            **row,
            'DECISION_USUARIO': '',
            'VALOR_USUARIO': '',
            'OBSERVACION_USUARIO': '',
            'PROPUESTA_EMPIRIA': proposal,
        })
    return result


def build_case_rows(rows):
    result = []
    for row in rows:
        result.append({
            **row,
            'DECISION_USUARIO': '',
            'VALOR_USUARIO': '',
            'OBSERVACION_USUARIO': '',
            'PROPUESTA_EMPIRIA': 'Diligencie exactamente VALOR en VALOR_USUARIO y MOTIVO + VIGENCIA_DESDE en OBSERVACION_USUARIO.',
        })
    return result


def build_location_rows(rows):
    result = []
    for row in rows:
        result.append({
            **row,
            'DECISION_USUARIO': '',
            'VALOR_USUARIO': '',
            'OBSERVACION_USUARIO': '',
            'PROPUESTA_EMPIRIA': 'Seleccione USAR_EXISTENTE, CREAR_NUEVA_UBICACION o NO_APLICA. Si usa existente, ponga el destino exacto en VALOR_USUARIO. Si crea nueva, ponga el nombre canonico en VALOR_USUARIO.',
        })
    return result


def build_catalog_rows(rows):
    result = []
    for row in rows:
        problema = clean(row.get('PROBLEMA'))
        proposal = clean(row.get('PROPUESTA_EMPIRIA'))
        if problema == 'TIPO_DOCUMENTO_PPT_NO_CATALOGADO':
            proposal = 'Si corresponde a PPT, seleccione CREAR_TIPO_PPT y conserve PERMISO POR PROTECCION TEMPORAL; no convertir a pasaporte.'
        elif problema == 'TIPO_VINCULACION_REQUERIDO':
            proposal = 'Seleccione exactamente LABORAL_OL, LABORAL_TF, LABORAL_TI u OPS.'
        else:
            proposal = 'Seleccione si acepta la opcion valida de focalizacion o si corrige la combinacion real de Personal.'
        result.append({
            **row,
            'DECISION_USUARIO': '',
            'VALOR_USUARIO': '',
            'OBSERVACION_USUARIO': '',
            'PROPUESTA_EMPIRIA': proposal,
        })
    return result


def apply_sheet_style(ws, title):
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    widths = [13, 17, 31, 34, 34, 70, 76, 28, 34, 46]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for cell in ws[1]:
        cell.fill = PatternFill('solid', fgColor=COLORS[title])
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(bottom=THIN)
        row[1].number_format = '@'
        for index in (7, 8, 9):
            if len(row) > index:
                row[index].fill = PatternFill('solid', fgColor=INPUT_FILL)


def add_sheet(workbook, title, rows, option_builder):
    ws = workbook.create_sheet(title)
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(header, '') for header in HEADERS])
    apply_sheet_style(ws, title)
    for excel_row, row in enumerate(rows, 2):
        options = option_builder(row)
        if options:
            validation = DataValidation(type='list', formula1='"' + ','.join(options) + '"', allow_blank=True)
            validation.error = 'Seleccione una opcion de la lista.'
            validation.showErrorMessage = True
            ws.add_data_validation(validation)
            validation.add(f'H{excel_row}')
    return ws


def option_fechas(_row):
    return ['PROPORCIONAR_FECHA', 'REVISAR_SOPORTE']


def option_identidades(row):
    if clean(row.get('PROBLEMA')) == 'CONFLICTO_IDENTIDAD':
        return ['MISMA_PERSONA', 'PERSONA_DISTINTA']
    return ['NOMBRE_LEGAL_COMPLETO']


def option_casos(_row):
    return ['COMPLETAR_CASO_ESPECIAL']


def option_ubicaciones(_row):
    return ['USAR_EXISTENTE', 'CREAR_NUEVA_UBICACION', 'NO_APLICA']


def option_catalogos(row):
    problema = clean(row.get('PROBLEMA'))
    if problema == 'TIPO_DOCUMENTO_PPT_NO_CATALOGADO':
        return ['CREAR_TIPO_PPT', 'CORREGIR_TIPO_DOCUMENTO']
    if problema == 'TIPO_VINCULACION_REQUERIDO':
        return ['LABORAL_OL', 'LABORAL_TF', 'LABORAL_TI', 'OPS']
    return ['ACEPTAR_OPCION_FOCALIZACION', 'CORREGIR_COMBINACION']


def add_summary(workbook, counts):
    ws = workbook.create_sheet('RESUMEN')
    ws.append(['INDICADOR', 'VALOR'])
    rows = [
        ('FECHA_CORTE', '2026-08-22'),
        ('RETIROS_2026_08_02_YA_ACEPTADOS', 17),
        ('DECISIONES_REALES_TOTAL', sum(counts.values())),
        ('FILAS_REALES_TOTAL', 32),
        ('FECHAS', counts['FECHAS']),
        ('IDENTIDADES', counts['IDENTIDADES']),
        ('CASOS_ESPECIALES', counts['CASOS_ESPECIALES']),
        ('UBICACIONES', counts['UBICACIONES']),
        ('CATALOGOS', counts['CATALOGOS']),
        ('OBJETIVO', 'Diligenciar unicamente estas 34 decisiones reales y volver a ejecutar el dry-run final.'),
    ]
    for item in rows:
        ws.append(item)
    apply_sheet_style(ws, 'RESUMEN')
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 80
    return ws


def main():
    wb = load_workbook(SOURCE)
    fechas = build_fecha_rows(read_sheet_rows(wb, 'FECHAS'))
    identidades = build_identity_rows(read_sheet_rows(wb, 'IDENTIDADES'))
    casos = build_case_rows(read_sheet_rows(wb, 'CASOS_ESPECIALES'))
    ubicaciones = build_location_rows(read_sheet_rows(wb, 'UBICACIONES_CARGOS'))
    catalogos = build_catalog_rows(read_sheet_rows(wb, 'CATALOGOS'))

    counts = {
        'FECHAS': len(fechas),
        'IDENTIDADES': len(identidades),
        'CASOS_ESPECIALES': len(casos),
        'UBICACIONES': len(ubicaciones),
        'CATALOGOS': len(catalogos),
    }
    if counts != {'FECHAS': 13, 'IDENTIDADES': 4, 'CASOS_ESPECIALES': 4, 'UBICACIONES': 9, 'CATALOGOS': 4}:
        raise RuntimeError(f'Conteo inesperado de decisiones reales: {counts}')

    out = Workbook()
    out.remove(out.active)
    add_sheet(out, 'FECHAS', fechas, option_fechas)
    add_sheet(out, 'IDENTIDADES', identidades, option_identidades)
    add_sheet(out, 'CASOS_ESPECIALES', casos, option_casos)
    add_sheet(out, 'UBICACIONES', ubicaciones, option_ubicaciones)
    add_sheet(out, 'CATALOGOS', catalogos, option_catalogos)
    add_summary(out, counts)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    out.save(OUTPUT)
    print(OUTPUT)


if __name__ == '__main__':
    main()

